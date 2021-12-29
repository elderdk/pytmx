from .constants import BASE_DIR, DEFAULT_KWARGS
from pathlib import Path


def get_outputtmxfilepath(filepath):
  tmx_filename = Path(filepath).stem + '.tmx'
  return BASE_DIR.joinpath(tmx_filename)


def get_complete_kwargs(**kwargs):
  return {**kwargs, **DEFAULT_KWARGS}


def show_progress(row):
    print(f'current row: {row}', end="\r")


def get_sh_maxrows_cell_value(filepath):
  suffix = Path(filepath).suffix

  if  suffix == '.xls':

    from xlrd import open_workbook

    wb = open_workbook(filepath)
    sh = wb.sheet_by_index(0)
    
    return (sh.nrows, sh.cell_value, 0, 0)

  elif suffix == '.xlsx':

    from openpyxl import load_workbook

    wb = load_workbook(filepath)
    sh = wb.active

    def _get_cell_value(row, col):
        return sh.cell(row, col).value

    return (sh.max_row, _get_cell_value, 1, 1)